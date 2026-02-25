import io
from datetime import datetime, date
from decimal import Decimal
from openpyxl import Workbook, load_workbook
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError
from ..models import TuitionPayment, StudentBalance, Salary, Expense
from apps.students.models import Student
from apps.university.models import Level, AcademicYear
from django.contrib.auth import get_user_model

User = get_user_model()


class PaymentExcelService:
    """Service to handle Excel import/export for tuition payments."""

    HEADERS = [
        'Matricule', 'Étudiant', 'Montant', 'Méthode de paiement',
        'Niveau', 'Date de paiement', 'Description', 'Référence', 'Statut'
    ]

    IMPORT_HEADERS = [
        'Matricule', 'Montant', 'Méthode de paiement',
        'Niveau', 'Date de paiement', 'Description'
    ]

    PAYMENT_METHODS = {
        'ESPÈCES': 'CASH', 'ESPECES': 'CASH', 'CASH': 'CASH',
        'VIREMENT': 'BANK_TRANSFER', 'VIREMENT BANCAIRE': 'BANK_TRANSFER', 'BANK_TRANSFER': 'BANK_TRANSFER',
        'MOBILE MONEY': 'MOBILE_MONEY', 'MOBILE_MONEY': 'MOBILE_MONEY',
        'CHÈQUE': 'CHECK', 'CHEQUE': 'CHECK', 'CHECK': 'CHECK',
    }

    @staticmethod
    def export_payments(queryset):
        """Export payments to an Excel workbook."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Paiements"

        ws.append(PaymentExcelService.HEADERS)

        for payment in queryset.select_related('student__user', 'level'):
            ws.append([
                payment.student.student_id,
                payment.student.user.get_full_name(),
                float(payment.amount),
                payment.get_payment_method_display(),
                payment.level.display_name if payment.level else '',
                payment.payment_date,
                payment.description,
                payment.reference,
                payment.get_status_display(),
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def download_template():
        """Generate a template Excel file for payment import."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Template Import Paiements"

        ws.append(PaymentExcelService.IMPORT_HEADERS)

        # Example row
        ws.append([
            "ETU2024001", 75000, "CASH",
            "L1", "2024-10-15", "Paiement 1er versement"
        ])

        # Add a second sheet with reference data
        ws_ref = wb.create_sheet("Référence")
        ws_ref.append(["Méthodes de paiement acceptées"])
        for display, code in [
            ("CASH ou ESPÈCES", "Espèces"),
            ("BANK_TRANSFER ou VIREMENT", "Virement bancaire"),
            ("MOBILE_MONEY", "Mobile Money"),
            ("CHECK ou CHÈQUE", "Chèque"),
        ]:
            ws_ref.append([display, code])

        ws_ref.append([])
        ws_ref.append(["Format date", "AAAA-MM-JJ (ex: 2024-10-15)"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def import_payments(file_obj, user):
        """
        Import payments from an Excel file.
        Returns (success_count, error_list)
        """
        try:
            wb = load_workbook(file_obj, data_only=True)
        except Exception as e:
            return 0, [f"Format de fichier invalide: {str(e)}"]

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        success_count = 0
        errors = []

        # Get current academic year
        academic_year = AcademicYear.objects.filter(is_current=True).first()
        if not academic_year:
            academic_year = AcademicYear.objects.order_by('-start_date').first()
        if not academic_year:
            return 0, ["Aucune année académique trouvée."]

        for row_idx, row in enumerate(rows, start=2):
            if not any(row):
                continue

            try:
                with transaction.atomic():
                    # Headers: Matricule[0], Montant[1], Méthode[2],
                    #          Niveau[3], Date[4], Description[5]

                    matricule = str(row[0]).strip() if row[0] else None
                    amount_val = row[1]
                    method_raw = str(row[2]).strip().upper() if row[2] else 'CASH'
                    level_name = str(row[3]).strip() if row[3] else None
                    date_val = row[4]
                    description = str(row[5]).strip() if row[5] and row[5] is not None else ''

                    if not matricule:
                        raise ValidationError("Le matricule est obligatoire.")

                    if not amount_val:
                        raise ValidationError("Le montant est obligatoire.")

                    # Parse amount
                    try:
                        amount = Decimal(str(amount_val))
                        if amount <= 0:
                            raise ValueError()
                    except (ValueError, Exception):
                        raise ValidationError(f"Montant invalide: {amount_val}")

                    # Find student
                    try:
                        student = Student.objects.get(student_id=matricule)
                    except Student.DoesNotExist:
                        raise ValidationError(f"Étudiant avec matricule '{matricule}' introuvable.")

                    # Parse payment method
                    payment_method = PaymentExcelService.PAYMENT_METHODS.get(method_raw, 'CASH')

                    # Parse level
                    level = None
                    if level_name:
                        level = Level.objects.filter(name__iexact=level_name).first()
                        if not level:
                            level = Level.objects.filter(name__icontains=level_name).first()

                    # Parse date
                    payment_date = None
                    if date_val:
                        if isinstance(date_val, (date, datetime)):
                            payment_date = date_val.date() if isinstance(date_val, datetime) else date_val
                        elif isinstance(date_val, str):
                            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                                try:
                                    payment_date = datetime.strptime(date_val, fmt).date()
                                    break
                                except ValueError:
                                    continue
                    if not payment_date:
                        payment_date = timezone.now().date()

                    # Generate reference
                    import uuid
                    reference = f"PAY-{timezone.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"

                    # Create payment
                    payment = TuitionPayment.objects.create(
                        student=student,
                        academic_year=academic_year,
                        level=level,
                        amount=amount,
                        payment_method=payment_method,
                        status='COMPLETED',
                        reference=reference,
                        description=description,
                        payment_date=payment_date,
                        received_by=user,
                    )

                    # Update student balance
                    balance, _ = StudentBalance.objects.get_or_create(
                        student=student,
                        academic_year=academic_year,
                        defaults={'total_due': 0, 'total_paid': 0}
                    )
                    balance.total_paid += amount
                    balance.save()

                    success_count += 1

            except Exception as e:
                errors.append(f"Ligne {row_idx}: {str(e)}")

        return success_count, errors


class SalaryExcelService:
    """Service to handle Excel import/export for salaries."""

    HEADERS = [
        'Employé', 'Email', 'Mois', 'Année', 'Salaire de base',
        'Primes', 'Déductions', 'Salaire net', 'Statut', 'Date de paiement'
    ]

    IMPORT_HEADERS = [
        'Email', 'Mois', 'Année', 'Salaire de base', 'Primes', 'Déductions'
    ]

    @staticmethod
    def export_salaries(queryset):
        """Export salaries to an Excel workbook."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Salaires"

        ws.append(SalaryExcelService.HEADERS)

        for salary in queryset.select_related('employee'):
            ws.append([
                salary.employee.get_full_name(),
                salary.employee.email,
                salary.month,
                salary.year,
                float(salary.base_salary),
                float(salary.bonuses),
                float(salary.deductions),
                float(salary.net_salary),
                salary.get_status_display(),
                salary.payment_date,
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def download_template():
        """Generate a template Excel file for salary import."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Template Import Salaires"

        ws.append(SalaryExcelService.IMPORT_HEADERS)

        ws.append([
            "jean.dupont@example.com", 1, 2025, 300000, 50000, 25000
        ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def import_salaries(file_obj, user):
        """Import salaries from an Excel file. Returns (success_count, error_list)"""
        try:
            wb = load_workbook(file_obj, data_only=True)
        except Exception as e:
            return 0, [f"Format de fichier invalide: {str(e)}"]

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        success_count = 0
        errors = []

        for row_idx, row in enumerate(rows, start=2):
            if not any(row):
                continue

            try:
                with transaction.atomic():
                    # Headers: Email[0], Mois[1], Année[2], Base[3], Primes[4], Déductions[5]
                    email = str(row[0]).strip() if row[0] else None
                    month_val = row[1]
                    year_val = row[2]
                    base_val = row[3]
                    bonuses_val = row[4] if row[4] else 0
                    deductions_val = row[5] if row[5] else 0

                    if not email:
                        raise ValidationError("L'email est obligatoire.")
                    if not month_val or not year_val or not base_val:
                        raise ValidationError("Mois, Année et Salaire de base sont obligatoires.")

                    try:
                        employee = User.objects.get(email=email)
                    except User.DoesNotExist:
                        raise ValidationError(f"Employé avec email '{email}' introuvable.")

                    month = int(month_val)
                    year = int(year_val)
                    base_salary = Decimal(str(base_val))
                    bonuses = Decimal(str(bonuses_val))
                    deductions = Decimal(str(deductions_val))
                    net_salary = base_salary + bonuses - deductions

                    if Salary.objects.filter(employee=employee, month=month, year=year).exists():
                        raise ValidationError(f"Salaire déjà enregistré pour {employee.get_full_name()} {month}/{year}.")

                    Salary.objects.create(
                        employee=employee,
                        month=month,
                        year=year,
                        base_salary=base_salary,
                        bonuses=bonuses,
                        deductions=deductions,
                        net_salary=net_salary,
                        status='PENDING',
                        processed_by=user,
                    )

                    success_count += 1

            except Exception as e:
                errors.append(f"Ligne {row_idx}: {str(e)}")

        return success_count, errors


class ExpenseExcelService:
    """Service to handle Excel import/export for expenses."""

    HEADERS = [
        'Date', 'Catégorie', 'Description', 'Montant', 'Créé par'
    ]

    IMPORT_HEADERS = [
        'Date', 'Catégorie', 'Description', 'Montant'
    ]

    CATEGORIES = {
        'SALAIRES': 'SALARIES', 'SALARIES': 'SALARIES',
        'SERVICES PUBLICS': 'UTILITIES', 'UTILITIES': 'UTILITIES',
        'MAINTENANCE': 'MAINTENANCE',
        'ÉQUIPEMENT': 'EQUIPMENT', 'EQUIPEMENT': 'EQUIPMENT', 'EQUIPMENT': 'EQUIPMENT',
        'FOURNITURES': 'SUPPLIES', 'SUPPLIES': 'SUPPLIES',
        'AUTRES': 'OTHER', 'OTHER': 'OTHER',
    }

    @staticmethod
    def export_expenses(queryset):
        """Export expenses to an Excel workbook."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Dépenses"

        ws.append(ExpenseExcelService.HEADERS)

        for expense in queryset.select_related('created_by'):
            ws.append([
                expense.date,
                expense.get_category_display(),
                expense.description,
                float(expense.amount),
                expense.created_by.get_full_name() if expense.created_by else '',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def download_template():
        """Generate a template Excel file for expense import."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Template Import Dépenses"

        ws.append(ExpenseExcelService.IMPORT_HEADERS)

        ws.append([
            "2025-01-15", "MAINTENANCE", "Réparation climatiseur", 150000
        ])

        ws_ref = wb.create_sheet("Référence")
        ws_ref.append(["Catégories acceptées"])
        for code, label in [
            ("SALARIES", "Salaires"),
            ("UTILITIES", "Services publics"),
            ("MAINTENANCE", "Maintenance"),
            ("EQUIPMENT", "Équipement"),
            ("SUPPLIES", "Fournitures"),
            ("OTHER", "Autres"),
        ]:
            ws_ref.append([code, label])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def import_expenses(file_obj, user):
        """Import expenses from an Excel file. Returns (success_count, error_list)"""
        try:
            wb = load_workbook(file_obj, data_only=True)
        except Exception as e:
            return 0, [f"Format de fichier invalide: {str(e)}"]

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        success_count = 0
        errors = []

        for row_idx, row in enumerate(rows, start=2):
            if not any(row):
                continue

            try:
                with transaction.atomic():
                    # Headers: Date[0], Catégorie[1], Description[2], Montant[3]
                    date_val = row[0]
                    category_raw = str(row[1]).strip().upper() if row[1] else None
                    description = str(row[2]).strip() if row[2] else ''
                    amount_val = row[3]

                    if not category_raw:
                        raise ValidationError("La catégorie est obligatoire.")
                    if not amount_val:
                        raise ValidationError("Le montant est obligatoire.")

                    category = ExpenseExcelService.CATEGORIES.get(category_raw)
                    if not category:
                        raise ValidationError(f"Catégorie invalide: {category_raw}. Valeurs: SALARIES, UTILITIES, MAINTENANCE, EQUIPMENT, SUPPLIES, OTHER")

                    amount = Decimal(str(amount_val))
                    if amount <= 0:
                        raise ValidationError(f"Montant invalide: {amount_val}")

                    # Parse date
                    expense_date = None
                    if date_val:
                        if isinstance(date_val, (date, datetime)):
                            expense_date = date_val.date() if isinstance(date_val, datetime) else date_val
                        elif isinstance(date_val, str):
                            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                                try:
                                    expense_date = datetime.strptime(date_val, fmt).date()
                                    break
                                except ValueError:
                                    continue
                    if not expense_date:
                        expense_date = timezone.now().date()

                    Expense.objects.create(
                        date=expense_date,
                        category=category,
                        description=description,
                        amount=amount,
                        created_by=user,
                    )

                    success_count += 1

            except Exception as e:
                errors.append(f"Ligne {row_idx}: {str(e)}")

        return success_count, errors

