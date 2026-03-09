import io
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.db import models, transaction
from django.core.exceptions import ValidationError
from ..models import Student, Enrollment
from apps.university.models import Program, Level, AcademicYear

User = get_user_model()

class StudentExcelService:
    """Service to handle Excel import/export for students."""

    HEADERS = [
        'Matricule', 'Prénom', 'Nom', 'Email', 'Sexe', 'Date Naissance',
        'Téléphone', 'Code Programme', 'Niveau Actuel', 'Date Inscription',
        'Statut', 'Nom Tuteur', 'Téléphone Tuteur', 'Contact Urgence'
    ]

    @staticmethod
    def export_students(queryset):
        """Export students to an Excel workbook."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Étudiants"

        # Add headers
        ws.append(StudentExcelService.HEADERS)

        # Add student data
        for student in queryset:
            ws.append([
                student.student_id,
                student.user.first_name,
                student.user.last_name,
                student.user.email,
                student.user.get_gender_display(),
                student.user.date_of_birth,
                student.user.phone,
                student.program.code if student.program else "",
                student.current_level.name if student.current_level else "",
                student.enrollment_date,
                student.get_status_display(),
                student.guardian_name,
                student.guardian_phone,
                student.emergency_contact
            ])

        # Write to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def download_template():
        """Generate a template Excel file for student import."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Template Import"

        # Add headers
        ws.append(StudentExcelService.HEADERS)

        # Add an example row
        ws.append([
            "", "Jean", "Dupont", "jean.dupont@example.com", "M", "2000-01-01",
            "+221770000000", "INFO",
            "L1", "2023-10-01", "ACTIVE", "Tuteur Dupont",
            "+221780000000", "Contact Urgence"
        ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def import_students(file_obj):
        """
        Import students from an Excel file.
        Returns (success_count, error_list)
        """
        try:
            wb = load_workbook(file_obj, data_only=True)
        except Exception as e:
            return 0, [f"Format de fichier invalide: {str(e)}"]

        ws = wb.active

        # --- Detect header row and build column index map ---
        header_row = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        def col(names):
            """Return first matching column index (0-based) from list of possible header names."""
            for name in names:
                for i, h in enumerate(header_row):
                    if h.lower() == name.lower():
                        return i
            return None

        idx_matricule   = col(['Matricule'])
        idx_first_name  = col(['Prénom', 'Prenom', 'First Name'])
        idx_last_name   = col(['Nom', 'Last Name'])
        idx_email       = col(['Email', 'E-mail'])
        idx_gender      = col(['Sexe', 'Genre', 'Gender'])
        idx_birth       = col(['Date Naissance', 'Date de naissance', 'Naissance'])
        idx_phone       = col(['Téléphone', 'Telephone', 'Phone'])
        idx_program     = col(['Code Programme', 'Programme', 'Program'])
        idx_level       = col(['Niveau Actuel', 'Niveau', 'Level'])
        idx_enroll_date = col(['Date Inscription', 'Date d\'inscription'])
        idx_status      = col(['Statut', 'Status'])
        idx_guardian    = col(['Nom Tuteur', 'Tuteur'])
        idx_guardian_ph = col(['Téléphone Tuteur'])
        idx_emergency   = col(['Contact Urgence', 'Urgence'])

        def get(row, idx, default=None):
            if idx is None or idx >= len(row):
                return default
            val = row[idx]
            return str(val).strip() if val is not None else default

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        success_count = 0
        errors = []

        for row_idx, row in enumerate(rows, start=2):
            if not any(row):  # Skip empty rows
                continue

            try:
                with transaction.atomic():

                    def parse_date(val):
                        if not val:
                            return None
                        if isinstance(val, (date, datetime)):
                            return val.date() if isinstance(val, datetime) else val
                        if isinstance(val, str):
                            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                                try:
                                    return datetime.strptime(val, fmt).date()
                                except ValueError:
                                    continue
                        return None

                    student_id_val   = get(row, idx_matricule)
                    first_name       = get(row, idx_first_name)
                    last_name        = get(row, idx_last_name)
                    email            = get(row, idx_email)
                    gender           = get(row, idx_gender) or 'M'
                    birth_date       = parse_date(row[idx_birth] if idx_birth is not None else None)
                    phone            = get(row, idx_phone) or ''
                    program_val      = get(row, idx_program)
                    level_name       = get(row, idx_level)
                    enroll_date_raw  = row[idx_enroll_date] if idx_enroll_date is not None else None
                    enroll_date      = parse_date(enroll_date_raw) or timezone.now().date()
                    status           = get(row, idx_status) or 'ACTIVE'
                    guardian_name    = get(row, idx_guardian) or ''
                    guardian_phone   = get(row, idx_guardian_ph) or ''
                    emergency_contact= get(row, idx_emergency) or ''

                    if not (first_name and last_name and program_val):
                        raise ValidationError("Les champs Prénom, Nom et Programme sont obligatoires.")

                    # Auto-generate email if missing
                    import re, unicodedata
                    def slugify(text):
                        text = unicodedata.normalize('NFKD', text)
                        text = text.encode('ascii', 'ignore').decode('ascii')
                        text = re.sub(r'[^\w\s-]', '', text).strip().lower()
                        return re.sub(r'[\s]+', '.', text)

                    if not email:
                        base = f"{slugify(first_name)}.{slugify(last_name)}"
                        if not base or base == '.':
                            # Fallback for non-latin names: use row index
                            base = f"etudiant{row_idx}"
                        
                        email = f"{base}@attawoune.edu"
                        # Make email unique
                        counter = 1
                        while User.objects.filter(email=email).exists():
                            email = f"{base}{counter}@attawoune.edu"
                            counter += 1
                    else:
                        # If email provided in excel but exists, error out
                        if User.objects.filter(email=email).exists():
                            raise ValidationError(f"Un utilisateur avec l'email {email} existe déjà.")

                    # Normalize gender
                    gender_upper = gender.upper() if gender else 'M'
                    if gender_upper in ('F', 'FEMME', 'FEMALE'):
                        gender_norm = 'F'
                    else:
                        gender_norm = 'M'

                    # Generate a predictable password like "FirstLast@YYYY"
                    birth_year = birth_date.year if getattr(birth_date, 'year', None) else date.today().year
                    generated_password = f"{first_name}{last_name}@{birth_year}".replace(" ", "")

                    # 1. Handle User
                    user = User.objects.create_user(
                        email=email,
                        username=email,
                        password=generated_password,
                        first_name=first_name,
                        last_name=last_name,
                        role='STUDENT',
                        gender=gender_norm,
                        date_of_birth=birth_date,
                        phone=phone,
                    )

                    # 2. Find Program — try code first, then full name
                    program_code_part = None
                    if " - " in program_val:
                        program_code_part = program_val.split(" - ")[0].strip()

                    program = None
                    if program_code_part:
                        program = Program.objects.filter(code__iexact=program_code_part).first()
                    if not program:
                        program = Program.objects.filter(code__iexact=program_val).first()
                    if not program:
                        program = Program.objects.filter(name__iexact=program_val).first()
                    if not program:
                        program = Program.objects.filter(name__icontains=program_val).first()

                    if not program:
                        raise ValidationError(f"Programme '{program_val}' introuvable. Vérifiez le nom ou le code du programme.")

                    # 3. Find Level
                    level = None
                    if level_name:
                        level = Level.objects.filter(
                            models.Q(name__iexact=level_name) | models.Q(name__icontains=level_name)
                        ).first()
                    if not level:
                        level = program.levels.first()

                    # 4. Handle Student Profile
                    student = Student.objects.create(
                        user=user,
                        program=program,
                        current_level=level,
                        enrollment_date=enroll_date,
                        status=status,
                        guardian_name=guardian_name,
                        guardian_phone=guardian_phone,
                        emergency_contact=emergency_contact,
                    )

                    if student_id_val:
                        student.student_id = student_id_val
                        student.save()

                    # 5. Handle Enrollment for current year
                    try:
                        current_year = AcademicYear.objects.get(is_current=True)
                        Enrollment.objects.get_or_create(
                            student=student,
                            academic_year=current_year,
                            defaults={
                                'program': program,
                                'level': level,
                                'status': 'ENROLLED'
                            }
                        )
                    except AcademicYear.DoesNotExist:
                        pass

                    success_count += 1

            except Exception as e:
                errors.append(f"Ligne {row_idx}: {str(e)}")

        return success_count, errors

